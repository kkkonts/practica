import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Alignment
import tools

def output(model, filename):
    dff = pd.read_excel(tools.NATIONAL_PROJECTS_PATH)

    # Group the data by the specified column
    grouped_data = dff.groupby(by=['Краткое наименование национального проекта'])

    output_filename = "output/" + filename + ".xlsx"

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Iterate over each group
        for name, group in grouped_data:
            df_res = pd.DataFrame(
                columns=['№ п/п', 'id категории', 'Категория', 'НП', 'id подкатегории', 'Подкатегория', 'id факта', 'Факт']
            )
            i = 1
            for idx, row in group.iterrows():
                search_result = model.search(row['Наименование мероприятия'])

                row_df = pd.DataFrame({
                    '№ п/п': [i],
                    'id категории': [np.nan],
                    'Категория': [np.nan],
                    'НП': [row['Наименование мероприятия']],
                    'id подкатегории': [np.nan],
                    'Подкатегория': [np.nan],
                    'id факта': [np.nan],
                    'Факт': [np.nan]
                })

                search_result_df = pd.DataFrame({
                    '№ п/п': np.nan,
                    'id категории': search_result['id категории'],
                    'Категория': search_result['категория'],
                    'НП': np.nan,
                    'id подкатегории': search_result['id подкатегории'],
                    'Подкатегория': search_result['подкатегория'],
                    'id факта': search_result['id факта'],
                    'Факт': search_result['факт']
                })

                df_res = pd.concat([df_res, row_df, search_result_df.dropna(axis=1, how='all')], ignore_index=True,
                                   axis=0)
                i += 1

            sheet_name = name[0]  # use the first character of the group name for the sheet name
            df_res.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            blue_fill = PatternFill(start_color="E0EAF6", end_color="E0EAF6", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            wrap_alignment = Alignment(wrap_text=True)

            # Find duplicates in 'id факта'
            id_fact_counts = df_res['id факта'].value_counts()
            duplicates = id_fact_counts[id_fact_counts > 1].index

            # Create a set of indices for rows to be highlighted
            highlight_rows = set()

            for idx, row in df_res.iterrows():
                if not pd.isna(row['НП']):
                    # Apply blue fill and wrap text to the merged cells
                    start_row = idx + 2
                    start_col = 2
                    end_row = idx + 2
                    end_col = 8
                    cell = worksheet.cell(row=start_row, column=start_col)
                    cell.value = row['НП']
                    worksheet.cell(row=start_row, column=1).fill = blue_fill
                    cell.alignment = wrap_alignment
                    worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

                # Highlight rows with duplicate 'id факта'
                if row['id факта'] in duplicates:
                    highlight_rows.add(idx + 2)  # Store 1-based index for highlighting

            for row_idx in highlight_rows:
                for col in range(1, len(df_res.columns) + 1):  # A to H columns (1-based index)
                    worksheet.cell(row=row_idx, column=col).fill = yellow_fill

            # Autofit column widths after applying text wrapping
            for col in worksheet.columns:
                max_length = 15
                column = col[0].column_letter  # Get the column name
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

    print(f"Output saved to {output_filename}")


